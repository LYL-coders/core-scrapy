import sqlite3

import requests
from bs4 import BeautifulSoup
import re
import openpyxl
import time
# import xlrd

from os import path
from PIL import Image
import numpy as np
import matplotlib.pyplot as plt
import jieba
from wordcloud import WordCloud, STOPWORDS

headers = {'user-agent':
             'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_3) '
             'AppleWebKit/537.36 (KHTML, like Gecko) '
             'Chrome/79.0.3945.79 Safari/537.36'}

class AttackCrawling_content:
    def __init__(self,headers):
        self.headers = headers
        self.str_all = ""

    def crwal(self,url):
        # get the content
        self.url = url
        res_news = requests.get(self.url,headers=self.headers)
        html_text = res_news.text
        bs = BeautifulSoup(html_text, 'html.parser')
        paras = bs.find_all('p')
        cur_str = ""
        for para in paras:
            content = "".join('% s' % id for id in para)  # 全部转为字符串
            cur_str += "".join(re.findall('[0-9\u4e00-\u9fa5]', content))
        self.str_all += cur_str

class AttackCrawling_link:
    def __init__(self,url,params,headers):
        self.url = url
        self.params = params
        self.headers = headers

    def crawl(self,n):
        # n表示爬取次数
        self.n = n
        self.all_link_title = []
        self.all_links = []
        for i in range(n):
            print("Now we are crawling the {0}th page".format(i+1))
            links_curr_page = []
            titles_curr_page = []
            res_news = requests.get(self.url,params=self.params,headers=self.headers)
            html_text = res_news.text
            bs = BeautifulSoup(html_text,'html.parser')
            a_links = bs.find_all('a',class_='news-title-font_1xS-F')
            for a_link in a_links:
                links_curr_page.append(a_link['href'])
                title = "".join('% s' % id for id in a_link.contents) # 全部转为字符串
                titles_curr_page.append("".join(re.findall('[\u4e00-\u9fa5]',title)))
            self.all_link_title += titles_curr_page
            self.all_links += links_curr_page
            self.params['pn'] += 10

    def writeToFile(self,dirPath,fileName):
        print("Now we are saving information")
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = time.strftime("%Y-%m-%d", time.localtime())
        sheet.append(['编号', '文章标题', '全文链接'])
        for i,(title,link) in enumerate(zip(self.all_link_title,self.all_links)):
            sheet.append([i+1,title,link])
        wb.save(dirPath + '/' + fileName)
        wb.close()

class WordCloudGenerator:
    def __init__(self, filePath=None, fontPath=None, wc_name=None, maskPath=None):
        d = path.dirname(__file__)  # 当前文件夹
        self.filePath = filePath
        try:
            self.file = open(path.join(d, self.filePath),encoding='gbk').read()
        except BaseException as e:
            print("we use the str in program")
            pass
        self.fontPath = fontPath
        self.maskPath = maskPath
        self.wc_name = wc_name
        try:
            self.mask = np.array(Image.open(path.join(d, self.maskPath)))
        except BaseException as e:
            print("use the defualt mask")
            self.mask = None

    def generateWordCloud(self):
        default_mode = jieba.cut(self.file)
        text = " ".join(default_mode)
        stopwords = set(STOPWORDS)
        stopwords.add("said")
        self.wc = WordCloud(
            # 设置字体，不指定就会出现乱码,这个字体文件需要下载
            font_path=self.fontPath,
            background_color="white",
            max_words=2000,
            mask=self.mask,
            stopwords=stopwords)
        self.wc.generate(text)

    def writeToImg(self):
        d = path.dirname(__file__)
        self.wc.to_file(path.join(d, self.wc_name))

    def show(self):
        plt.imshow(self.wc, interpolation='bilinear')
        plt.axis("off")
        plt.figure()




if __name__ =="__main__":
    needCraw = False
    filePath = r'D:\core-scrapy\beautifulsoup\data\crawling_data\attack_news\新闻列表.xlsx'
    url = 'https://www.baidu.com/s'
    params = {'ie': 'utf-8', 'medium': 0, 'rtt': 1, 'bsst': 1, 'rsv_dl': 'news_t_sk', 'cl': 2, 'wd': '爬虫技术',
              'tn': 'news', 'rsv_bp': 1, 'tfflag': 0, 'x_bfe_rqs': '03E80000001', 'x_bfe_tjscore': '0.100000',
              'tngroupname': 'organic_news', 'newVideo': 12, 'pn': 0}
    headers = {'user-agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_14_3) '
                             'AppleWebKit/537.36 (KHTML, like Gecko) '
                             'Chrome/79.0.3945.79 Safari/537.36'}
    try:
        data = []
        wb = openpyxl.load_workbook(filePath)
        print("加载成功")
        table = wb.get_sheet_by_name("2022-09-25")
        n_rows = table.max_row
        for row in range(2,n_rows):
            data.append({"title":table.cell(row,2).value,"link":table.cell(row,3).value})

        # 按照标题产生词云
        link_title= []
        str_title = ""
        for item in data:
            link_title.append(item['link'])
        str_title = " ".join(link_title)
        # //wcg_title = WordCloudGenerator(fontPath='msyh.ttc', wc_name="titleWordCloud.tif")
        # //wcg_title.file = str_title
        # //wcg_title.generateWordCloud()
        # //wcg_title.writeToImg()
        # //wcg_title.show()

        #按照内容产生词云
        attCrawlCont = AttackCrawling_content(headers=headers)
        for i,link in enumerate(link_title):
            print("Now we are crawling content in {0}th page".format(i+1))
            attCrawlCont.crwal(url=link)
            with open('all_files.txt','w+') as f:
                f.write(attCrawlCont.str_all)

        wcg_content = WordCloudGenerator(filePath='all_files.txt',fontPath='msyh.ttc',wc_name="contentWordCloud.tif")
        wcg_content.generateWordCloud()
        wcg_content.writeToImg()
        wcg_content.show()

    except BaseException as e:
        print("The file does not exist, so we begin crawling")
        print(e)
        attackCraw = AttackCrawling_link(url=url, params=params, headers=headers)
        dirPath = r'D:\core-scrapy\beautifulsoup\data\crawling_data\attack_news'
        attackCraw.crawl(10)
        attackCraw.writeToFile(dirPath=dirPath,fileName="新闻列表.xlsx")
