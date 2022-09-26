import xlrd
import pymysql
import xlsxwriter
import openpyxl

# import importlib
# importlib.reload(sys) #出现呢reload错误使用
def open_excel():
    try:
        # book = xlrd.open_workbook(r"D:\core-scrapy\beautifulsoup\data\crawling_data\attack_news\新闻列表.xlsx")  # 文件名，把文件与py文件放在同一目录下
        # book = xlrd.open_workbook_xls(r"D:\core-scrapy\beautifulsoup\data\crawling_data\attack_news\news.xlsx")
        # book = xlsxwriter.Workbook(r"D:\core-scrapy\beautifulsoup\data\crawling_data\attack_news\news.xlsx")
        book = openpyxl.load_workbook(r"D:\core-scrapy\beautifulsoup\data\crawling_data\attack_news\news.xlsx")

    except Exception as e:
        print(e)
        print("open excel file failed!")
    try:
        # sheet = book.sheet_by_name("sheet名称")  # execl里面的worksheet1
        # sheet = book.sheet_by_name('2022-09-25')
        sheet = book['2022-09-25']
        return sheet
    except:
        print("locate worksheet in excel failed!")
        # 连接数据库

 # self.con = pymysql.connect(host="127.0.0.1", port=3306, user="root", passwd="123456", charset="utf8")
 #            self.cursor = self.con.cursor(pymysql.cursors.DictCursor)
try:
    db = pymysql.connect(host="127.0.0.1",port=3306, user="root",
                         passwd="123456",
                         db="xlsx",
                         charset='utf8')
except Exception as e:
    print("could not connect to mysql server")
    print(e)


def search_count():
    cursor = db.cursor()
    select = "select count(id) from XXXX"  # 获取表中xxxxx记录数
    cursor.execute(select)  # 执行sql语句
    line_count = cursor.fetchone()
    print(line_count[0])


def insert_deta():
    sheet = open_excel()
    # print('sheet.columns:::',list(sheet.columns))
    cursor = db.cursor()
    for i in range(2, 50):  # 第一行是标题名，对应表中的字段名所以应该从第二行开始，计算机以0开始计数，所以值是1

        name = sheet.cell(i, 1).value  # 取第i行第0列
        data = sheet.cell(i, 2).value  # 取第i行第1列，下面依次类推
        data2 = sheet.cell(i, 3).value
        print(name)
        print(data)
        value = (name, data, data2)
        print(value)
        sql = "INSERT INTO news(name,data,data2)VALUES(%s,%s,%s)"
        cursor.execute(sql, value)  # 执行sql语句
        db.commit()
    cursor.close()  # 关闭连接


insert_deta()
db.close()  # 关闭数据
print("ok ")